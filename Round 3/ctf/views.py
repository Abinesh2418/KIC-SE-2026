import os
import zipfile
import io
from functools import wraps

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout, get_user_model
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse, Http404
from django.views.decorators.http import require_POST, require_GET
from django.views.decorators.csrf import csrf_exempt
from django.utils import timezone
from django.conf import settings
from django.db.models import Sum, F, Q, Count

from .models import (
    SiteSettings, Challenge, Flag, ChallengeResource,
    Submission, Score, EvaluatorAssignment, UserFlag
)
from .forms import RegisterForm, LoginForm

User = get_user_model()


# ── Decorators ──────────────────────────────────────────────────

def admin_required(view_func):
    """Only allow admin users."""
    @wraps(view_func)
    @login_required
    def wrapper(request, *args, **kwargs):
        if not request.user.is_admin_user:
            messages.error(request, "Admin access required.")
            return redirect('home')
        return view_func(request, *args, **kwargs)
    return wrapper


def evaluator_required(view_func):
    """Only allow evaluator or admin users."""
    @wraps(view_func)
    @login_required
    def wrapper(request, *args, **kwargs):
        if not request.user.is_evaluator_user:
            messages.error(request, "Evaluator access required.")
            return redirect('home')
        return view_func(request, *args, **kwargs)
    return wrapper


# ── Notebook Auth API ───────────────────────────────────────────

@csrf_exempt
@require_POST
def notebook_login(request):
    """JSON-friendly login endpoint for Jupyter notebooks.
    Accepts POST with username & password, returns session cookie."""
    username = request.POST.get("username", "").strip()
    password = request.POST.get("password", "")

    if not username or not password:
        return JsonResponse({"success": False, "message": "Username and password are required."}, status=400)

    user = authenticate(request, username=username, password=password)
    if user is None:
        return JsonResponse({"success": False, "message": "Invalid username or password."}, status=401)

    if not user.is_approved:
        return JsonResponse({"success": False, "message": "Your account is not yet approved."}, status=403)

    login(request, user)
    return JsonResponse({"success": True, "message": f"Logged in as {user.username}."})


# ── Home ────────────────────────────────────────────────────────

def home(request):
    return render(request, 'ctf/index.html')


# ── Auth ────────────────────────────────────────────────────────

def register_view(request):
    if request.user.is_authenticated:
        return redirect('challenges_list')

    if request.method == 'POST':
        # Handle both JSON and form-encoded
        if request.content_type == 'application/json':
            import json
            data = json.loads(request.body)
            form = RegisterForm(data)
        else:
            form = RegisterForm(request.POST)

        if form.is_valid():
            form.save()
            if request.content_type == 'application/json':
                return JsonResponse({'success': True, 'message': 'Account created! Please wait for admin approval.'})
            messages.success(request, 'Account created! Please wait for admin approval.')
            return redirect('login')
        else:
            if request.content_type == 'application/json':
                errors = '; '.join([f"{k}: {v[0]}" for k, v in form.errors.items()])
                return JsonResponse({'success': False, 'message': errors})
    else:
        form = RegisterForm()

    return render(request, 'ctf/register.html', {'form': form})


def login_view(request):
    if request.user.is_authenticated:
        return redirect('challenges_list')

    if request.method == 'POST':
        # Handle both JSON and form-encoded
        if request.content_type == 'application/json':
            import json
            data = json.loads(request.body)
            username = data.get('username', '')
            password = data.get('password', '')
        else:
            username = request.POST.get('username', '')
            password = request.POST.get('password', '')

        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            if request.content_type == 'application/json':
                return JsonResponse({'success': True})
            return redirect('challenges_list')
        else:
            if request.content_type == 'application/json':
                return JsonResponse({'success': False, 'message': 'Invalid username or password.'})
            messages.error(request, 'Invalid username or password.')

    return render(request, 'ctf/login.html')


def logout_view(request):
    logout(request)
    return redirect('home')


@login_required
def pending_approval(request):
    if request.user.is_approved:
        return redirect('challenges_list')
    return render(request, 'ctf/pending_approval.html')


# ── Challenges ──────────────────────────────────────────────────

@login_required
def challenges_list(request):
    site = SiteSettings.get()
    if request.user.is_admin_user or request.user.is_evaluator_user:
        challenges = Challenge.objects.all()
    else:
        challenges = Challenge.objects.filter(is_revealed=True)

    challenge_data = []
    for ch in challenges:
        flags = ch.flags.all()
        total_flags = flags.count()
        if request.user.is_authenticated:
            completed = Score.objects.filter(
                user=request.user, challenge=ch, is_approved=True
            ).count()
            pending = Score.objects.filter(
                user=request.user, challenge=ch, is_approved=False
            ).count()
            earned = Score.objects.filter(
                user=request.user, challenge=ch, is_approved=True
            ).aggregate(
                total=Sum(F('flag_points') + F('explanation_points'))
            )['total'] or 0
        else:
            completed = pending = earned = 0

        challenge_data.append({
            'challenge': ch,
            'total_flags': total_flags,
            'completed_flags': completed,
            'pending_flags': pending,
            'points_earned': earned,
        })

    return render(request, 'ctf/challenges.html', {
        'challenge_data': challenge_data,
        'site': site,
    })


@login_required
def challenge_detail(request, challenge_id):
    challenge = get_object_or_404(Challenge, pk=challenge_id)
    site = SiteSettings.get()

    # Non-staff can only see revealed challenges
    if not (request.user.is_admin_user or request.user.is_evaluator_user):
        if not challenge.is_revealed:
            raise Http404

    flags = challenge.flags.all()
    resources = challenge.resources.all()
    total_flags = flags.count()

    # Compute progress
    completed = Score.objects.filter(
        user=request.user, challenge=challenge, is_approved=True
    ).count()
    pending = Score.objects.filter(
        user=request.user, challenge=challenge, is_approved=False
    ).count()
    earned = Score.objects.filter(
        user=request.user, challenge=challenge, is_approved=True
    ).aggregate(
        total=Sum(F('flag_points') + F('explanation_points'))
    )['total'] or 0

    # Check which flags already have a score (pending or approved)
    existing_scores = {
        s.flag_id: s for s in Score.objects.filter(user=request.user, challenge=challenge)
    }

    flag_data = []
    for flag in flags:
        score = existing_scores.get(flag.id)
        status = None
        if score:
            status = 'approved' if score.is_approved else 'pending'
        flag_data.append({
            'flag': flag,
            'status': status,
            'score': score,
        })

    progress = {
        'completed_flags': completed,
        'pending_flags': pending,
        'total_flags': total_flags,
        'points_earned': earned,
        'total_possible': challenge.total_points,
    }

    return render(request, 'ctf/challenge_detail.html', {
        'challenge': challenge,
        'flag_data': flag_data,
        'resources': resources,
        'progress': progress,
        'event_active': site.event_active,
    })


@login_required
def challenge_download(request, challenge_id, filename):
    """Serve challenge files as a zip download."""
    challenge = get_object_or_404(Challenge, pk=challenge_id)

    if not request.user.is_approved and not request.user.is_admin_user:
        return HttpResponse("Account not approved.", status=403)

    # Collect all resource files for this challenge
    resources = challenge.resources.all()
    challenge_files_dir = settings.CHALLENGE_FILES_DIR

    if filename == 'all':
        # Create a zip of all challenge files
        buffer = io.BytesIO()
        with zipfile.ZipFile(buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
            for res in resources:
                filepath = os.path.join(challenge_files_dir, res.local_name)
                if os.path.exists(filepath):
                    zf.write(filepath, res.display_name)
        buffer.seek(0)
        safe_title = challenge.title.replace(' ', '_').replace(':', '')
        response = HttpResponse(buffer.read(), content_type='application/zip')
        response['Content-Disposition'] = f'attachment; filename="{safe_title}.zip"'
        return response
    else:
        # Serve individual file
        resource = challenge.resources.filter(local_name=filename).first()
        if not resource:
            raise Http404
        filepath = os.path.join(challenge_files_dir, resource.local_name)
        if not os.path.exists(filepath):
            raise Http404
        with open(filepath, 'rb') as f:
            content = f.read()

        # Determine content type
        if filename.endswith('.csv'):
            ct = 'text/csv'
        elif filename.endswith('.ipynb'):
            ct = 'application/json'
        elif filename.endswith('.pkl'):
            ct = 'application/octet-stream'
        else:
            ct = 'application/octet-stream'

        response = HttpResponse(content, content_type=ct)
        response['Content-Disposition'] = f'attachment; filename="{resource.display_name}"'
        return response


# ── Dashboard ───────────────────────────────────────────────────

@login_required
def dashboard(request):
    challenges = Challenge.objects.all()
    if not (request.user.is_admin_user or request.user.is_evaluator_user):
        challenges = challenges.filter(is_revealed=True)

    progress = {}
    total_approved_flags = 0
    total_earned_points = 0
    completed_challenges = 0
    total_challenges = challenges.count()

    for ch in challenges:
        flags = ch.flags.all()
        total_flags = flags.count()
        completed = Score.objects.filter(
            user=request.user, challenge=ch, is_approved=True
        ).count()
        pending = Score.objects.filter(
            user=request.user, challenge=ch, is_approved=False
        ).count()
        earned = Score.objects.filter(
            user=request.user, challenge=ch, is_approved=True
        ).aggregate(
            total=Sum(F('flag_points') + F('explanation_points'))
        )['total'] or 0

        progress[ch.id] = {
            'completed_flags': completed,
            'pending_flags': pending,
            'total_flags': total_flags,
            'points_earned': earned,
            'total_possible': ch.total_points,
        }

        total_approved_flags += completed
        total_earned_points += earned
        if total_flags > 0 and completed >= total_flags:
            completed_challenges += 1

    overall_pct = round((completed_challenges / total_challenges) * 100) if total_challenges > 0 else 0

    return render(request, 'ctf/dashboard.html', {
        'challenges': challenges,
        'progress': progress,
        'stats': {
            'total_points': total_earned_points,
            'approved_flags': total_approved_flags,
            'completed_challenges': completed_challenges,
            'total_challenges': total_challenges,
            'overall_pct': overall_pct,
        },
    })


# ── Flag Submission (AJAX) ──────────────────────────────────────

@require_POST
@login_required
def submit_flag(request, challenge_id, flag_order):
    import json
    site = SiteSettings.get()

    if not site.event_active:
        return JsonResponse({'success': False, 'message': 'Event is not active. Submissions are closed.'})

    if not request.user.is_approved:
        return JsonResponse({'success': False, 'message': 'Your account is not yet approved.'})

    challenge = get_object_or_404(Challenge, pk=challenge_id)
    flag = get_object_or_404(Flag, challenge=challenge, flag_order=flag_order)

    # Check if already has a score entry
    existing = Score.objects.filter(user=request.user, flag=flag).first()
    if existing:
        if existing.is_approved:
            return JsonResponse({'success': False, 'message': 'This flag has already been approved.'})
        else:
            return JsonResponse({'success': False, 'message': 'This flag is already pending evaluation.'})

    try:
        data = json.loads(request.body)
        submitted_value = data.get('flag', '').strip()
    except (json.JSONDecodeError, ValueError):
        submitted_value = request.POST.get('flag', '').strip()

    if not submitted_value:
        return JsonResponse({'success': False, 'message': 'Please enter a flag value.'})

    # Validate against the user's unique per-user flag first,
    # falling back to the global flag for backwards compatibility.
    from .models import UserFlag
    user_flag = UserFlag.objects.filter(user=request.user, challenge=challenge).first()
    if user_flag:
        is_correct = (submitted_value == user_flag.flag_value)
    else:
        is_correct = (submitted_value == flag.flag_content)
    Submission.objects.create(
        user=request.user,
        challenge=challenge,
        flag=flag,
        submitted_value=submitted_value,
        is_correct=is_correct,
    )

    if is_correct:
        # Create pending score
        Score.objects.create(
            user=request.user,
            challenge=challenge,
            flag=flag,
            flag_points=0,
            explanation_points=0,
            is_approved=False,
        )
        return JsonResponse({
            'success': True,
            'message': 'Correct flag! Submitted for evaluator approval.',
        })
    else:
        return JsonResponse({'success': False, 'message': 'Incorrect flag. Try again.'})


# ── Leaderboard ─────────────────────────────────────────────────

@login_required
def leaderboard_page(request):
    return render(request, 'ctf/leaderboard.html')


def leaderboard_api(request):
    site = SiteSettings.get()

    # Admin/evaluator always see leaderboard; participants only if public
    if request.user.is_authenticated and (request.user.is_admin_user or request.user.is_evaluator_user):
        pass  # always allowed
    elif not site.leaderboard_public:
        return JsonResponse({'success': False, 'message': 'Leaderboard is private'})

    # Get all approved participants with points.
    # Tiebreaker: among equal points the participant who reached that total
    # FIRST (earlier last_flag_at) ranks higher.  NULLs (no approved flags)
    # are pushed to the bottom.
    import datetime as _dt
    from django.db.models import Value
    from django.db.models.functions import Coalesce
    far_future = timezone.datetime(9999, 1, 1, tzinfo=_dt.timezone.utc)
    users = User.objects.filter(
        is_approved=True,
        role=User.ROLE_PARTICIPANT,
    ).annotate(
        sort_time=Coalesce('last_flag_at', Value(far_future)),
    ).order_by('-total_points', 'sort_time')

    leaderboard = []
    for i, user in enumerate(users, 1):
        challenges_completed = Score.objects.filter(
            user=user, is_approved=True
        ).values('challenge').distinct().count()

        leaderboard.append({
            'rank': i,
            'username': user.username,
            'total_points': user.total_points,
            'challenges_completed': challenges_completed,
            'last_flag_at': user.last_flag_at.isoformat() if user.last_flag_at else None,
        })

    return JsonResponse({'success': True, 'leaderboard': leaderboard})


@login_required
def export_leaderboard_xlsx(request):
    """Export leaderboard results as XLSX — admin and evaluators only."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    import datetime as _dt
    from django.db.models import Value
    from django.db.models.functions import Coalesce

    if not (request.user.is_admin_user or request.user.is_evaluator_user):
        return redirect('leaderboard')

    far_future = timezone.datetime(9999, 1, 1, tzinfo=_dt.timezone.utc)
    users = User.objects.filter(
        is_approved=True,
        role=User.ROLE_PARTICIPANT,
    ).annotate(
        sort_time=Coalesce('last_flag_at', Value(far_future)),
    ).order_by('-total_points', 'sort_time')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Round 3 - Leaderboard"

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="7B1FA2", end_color="7B1FA2", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    headers = ["Rank", "Username", "Email", "Total Points", "Challenges Completed", "Last Flag At"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for i, user in enumerate(users, 1):
        challenges_completed = Score.objects.filter(
            user=user, is_approved=True
        ).values('challenge').distinct().count()

        last_flag_str = user.last_flag_at.strftime("%Y-%m-%d %H:%M:%S") if user.last_flag_at else "—"

        row = [
            i,
            user.username,
            user.email,
            user.total_points,
            challenges_completed,
            last_flag_str,
        ]
        for col, val in enumerate(row, 1):
            cell = ws.cell(row=i + 1, column=col, value=val)
            cell.border = thin_border
            if col in (1, 4, 5):
                cell.alignment = Alignment(horizontal="center")

    # Auto-width columns
    for col_cells in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col_cells) + 2
        ws.column_dimensions[col_cells[0].column_letter].width = max_len

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="Round3_Leaderboard.xlsx"'
    wb.save(response)
    return response


# ── Admin Dashboard ─────────────────────────────────────────────

@admin_required
def admin_dashboard(request):
    site = SiteSettings.get()
    users = User.objects.all().order_by('-created_at')
    pending_users = User.objects.filter(is_approved=False).exclude(role=User.ROLE_ADMIN)
    challenges = Challenge.objects.all()
    evaluators = User.objects.filter(Q(role=User.ROLE_EVALUATOR) | Q(role=User.ROLE_ADMIN))
    participants = User.objects.filter(role=User.ROLE_PARTICIPANT, is_approved=True)
    assignments = EvaluatorAssignment.objects.select_related('evaluator', 'participant').all()

    # Pending scores for admin review
    pending_scores = Score.objects.filter(is_approved=False).select_related(
        'user', 'challenge', 'flag'
    )

    pending_rows = []
    for score in pending_scores:
        latest_sub = Submission.objects.filter(
            user=score.user, flag=score.flag, is_correct=True
        ).order_by('-submitted_at').first()

        pending_rows.append({
            'score': score,
            'latest_flag_value': latest_sub.submitted_value if latest_sub else '—',
            'submitted_at': latest_sub.submitted_at if latest_sub else score.submitted_at,
            'flag_points_max': score.challenge.flag_points_max,
            'explanation_points_max': score.challenge.explanation_points_max,
        })

    total_approved_scores = Score.objects.filter(is_approved=True).count()

    return render(request, 'ctf/admin_dashboard.html', {
        'site': site,
        'users': users,
        'pending_users': pending_users,
        'challenges': challenges,
        'evaluators': evaluators,
        'participants': participants,
        'assignments': assignments,
        'pending_rows': pending_rows,
        'total_approved_scores': total_approved_scores,
    })


@admin_required
@require_POST
def admin_toggle_event(request):
    site = SiteSettings.get()
    site.event_active = not site.event_active
    site.save()
    state = "started" if site.event_active else "stopped"
    messages.success(request, f"Event {state}.")
    return redirect('admin_dashboard')


@admin_required
@require_POST
def admin_toggle_leaderboard(request):
    site = SiteSettings.get()
    site.leaderboard_public = not site.leaderboard_public
    site.save()
    state = "public" if site.leaderboard_public else "private"
    messages.success(request, f"Leaderboard set to {state}.")
    return redirect('admin_dashboard')


@admin_required
@require_POST
def admin_reset_leaderboard(request):
    """Wipe all scores, submissions, and zero every participant's points
    so the challenge can be re-conducted from scratch."""
    # Delete all scores and submissions
    Score.objects.all().delete()
    Submission.objects.all().delete()
    # Zero out every participant's points and clear last_flag_at
    User.objects.filter(role=User.ROLE_PARTICIPANT).update(
        total_points=0,
        last_flag_at=None,
    )
    messages.success(request, "Leaderboard reset — all scores, submissions, and points have been cleared.")
    return redirect('admin_dashboard')


@admin_required
@require_POST
def admin_toggle_challenges(request):
    site = SiteSettings.get()
    site.challenges_revealed = not site.challenges_revealed
    site.save()
    # Also update all individual challenges
    Challenge.objects.all().update(is_revealed=site.challenges_revealed)
    state = "revealed" if site.challenges_revealed else "hidden"
    messages.success(request, f"All challenges {state}.")
    return redirect('admin_dashboard')


@admin_required
@require_POST
def admin_approve_user(request, user_id):
    user = get_object_or_404(User, pk=user_id)
    user.is_approved = True
    user.save(update_fields=['is_approved'])
    messages.success(request, f"User '{user.username}' approved.")
    return redirect('admin_dashboard')


@admin_required
@require_POST
def admin_reject_user(request, user_id):
    user = get_object_or_404(User, pk=user_id)
    username = user.username
    user.delete()
    messages.success(request, f"User '{username}' rejected and deleted.")
    return redirect('admin_dashboard')


@admin_required
@require_POST
def admin_delete_user(request, user_id):
    user = get_object_or_404(User, pk=user_id)
    if user.id == request.user.id:
        messages.error(request, "Cannot delete yourself.")
        return redirect('admin_dashboard')
    username = user.username
    user.delete()
    messages.success(request, f"User '{username}' deleted.")
    return redirect('admin_dashboard')


@admin_required
@require_POST
def admin_toggle_evaluator(request, user_id):
    user = get_object_or_404(User, pk=user_id)
    if user.is_admin_user:
        messages.error(request, "Cannot change admin roles.")
        return redirect('admin_dashboard')

    if user.role == User.ROLE_EVALUATOR:
        user.role = User.ROLE_PARTICIPANT
        messages.success(request, f"'{user.username}' demoted to participant.")
    else:
        user.role = User.ROLE_EVALUATOR
        user.is_approved = True
        messages.success(request, f"'{user.username}' promoted to evaluator.")
    user.save(update_fields=['role', 'is_approved'])
    return redirect('admin_dashboard')


@admin_required
@require_POST
def admin_assign_evaluator(request):
    evaluator_id = request.POST.get('evaluator_id')
    participant_id = request.POST.get('participant_id')

    if not evaluator_id or not participant_id:
        messages.error(request, "Both evaluator and participant must be selected.")
        return redirect('admin_dashboard')

    evaluator = get_object_or_404(User, pk=evaluator_id)
    participant = get_object_or_404(User, pk=participant_id)

    # Upsert: update if participant already assigned, else create
    assignment, created = EvaluatorAssignment.objects.update_or_create(
        participant=participant,
        defaults={'evaluator': evaluator},
    )
    if created:
        messages.success(request, f"Assigned '{evaluator.username}' to evaluate '{participant.username}'.")
    else:
        messages.success(request, f"Reassigned '{participant.username}' to evaluator '{evaluator.username}'.")
    return redirect('admin_dashboard')


@admin_required
@require_POST
def admin_unassign_evaluator(request, assignment_id):
    assignment = get_object_or_404(EvaluatorAssignment, pk=assignment_id)
    assignment.delete()
    messages.success(request, "Assignment removed.")
    return redirect('admin_dashboard')


@admin_required
@require_POST
def admin_toggle_challenge_reveal(request, challenge_id):
    challenge = get_object_or_404(Challenge, pk=challenge_id)
    challenge.is_revealed = not challenge.is_revealed
    challenge.save(update_fields=['is_revealed'])
    state = "revealed" if challenge.is_revealed else "hidden"
    messages.success(request, f"Challenge '{challenge.title}' {state}.")
    return redirect('admin_dashboard')


@admin_required
@require_POST
def admin_reveal_all(request):
    Challenge.objects.all().update(is_revealed=True)
    messages.success(request, "All challenges revealed.")
    return redirect('admin_dashboard')


@admin_required
@require_POST
def admin_hide_all(request):
    Challenge.objects.all().update(is_revealed=False)
    messages.success(request, "All challenges hidden.")
    return redirect('admin_dashboard')


@admin_required
@require_POST
def admin_update_challenge(request, challenge_id):
    challenge = get_object_or_404(Challenge, pk=challenge_id)
    challenge.title = request.POST.get('title', challenge.title)
    challenge.category = request.POST.get('category', challenge.category)
    challenge.difficulty = request.POST.get('difficulty', challenge.difficulty)
    tp = request.POST.get('total_points')
    if tp:
        challenge.total_points = int(tp)
    desc = request.POST.get('description')
    if desc:
        challenge.description = desc
    challenge.save()

    # Update flag content if provided
    flag_content = request.POST.get('flag_content')
    if flag_content:
        flag = challenge.flags.first()
        if flag:
            flag.flag_content = flag_content
            flag.save(update_fields=['flag_content'])

    messages.success(request, f"Challenge '{challenge.title}' updated.")
    return redirect('admin_dashboard')


# ── Evaluator Dashboard ─────────────────────────────────────────

@evaluator_required
def evaluator_dashboard(request):
    # Get participants assigned to this evaluator
    if request.user.is_admin_user:
        # Admin sees all pending scores
        assigned_ids = User.objects.filter(role=User.ROLE_PARTICIPANT).values_list('id', flat=True)
    else:
        assigned_ids = EvaluatorAssignment.objects.filter(
            evaluator=request.user
        ).values_list('participant_id', flat=True)

    pending_scores = Score.objects.filter(
        is_approved=False, user_id__in=assigned_ids
    ).select_related('user', 'challenge', 'flag')

    pending_rows = []
    for score in pending_scores:
        latest_sub = Submission.objects.filter(
            user=score.user, flag=score.flag, is_correct=True
        ).order_by('-submitted_at').first()

        pending_rows.append({
            'score': score,
            'latest_flag_value': latest_sub.submitted_value if latest_sub else '—',
            'submitted_at': latest_sub.submitted_at if latest_sub else score.submitted_at,
            'flag_points_max': score.challenge.flag_points_max,
            'explanation_points_max': score.challenge.explanation_points_max,
        })

    # Recently approved by this evaluator
    recent_approved = Score.objects.filter(
        is_approved=True, approved_by=request.user
    ).select_related('user', 'challenge', 'approved_by').order_by('-approved_at')[:20]

    return render(request, 'ctf/evaluator_dashboard.html', {
        'pending_rows': pending_rows,
        'recent_approved': recent_approved,
    })


@evaluator_required
@require_POST
def evaluator_approve(request, score_id):
    score = get_object_or_404(Score, pk=score_id, is_approved=False)

    flag_points = int(request.POST.get('flag_points', 0))
    explanation_points = int(request.POST.get('explanation_points', 0))

    # Cap to challenge max
    flag_points = min(flag_points, score.challenge.flag_points_max)
    explanation_points = min(explanation_points, score.challenge.explanation_points_max)

    score.flag_points = flag_points
    score.explanation_points = explanation_points
    score.is_approved = True
    score.approved_by = request.user
    score.approved_at = timezone.now()
    score.save()

    # Recalculate user points
    score.user.recalculate_points()

    messages.success(request, f"Score approved: {flag_points} flag + {explanation_points} explanation pts for {score.user.username}.")
    return redirect('evaluator_dashboard')


@evaluator_required
@require_POST
def evaluator_reject(request, score_id):
    score = get_object_or_404(Score, pk=score_id, is_approved=False)
    username = score.user.username
    score.delete()
    messages.success(request, f"Submission by '{username}' rejected.")
    return redirect('evaluator_dashboard')


# ── Admin: View All User Flags ───────────────────────────────────

@admin_required
def admin_user_flags(request):
    """Show every participant's 5 unique flags in a table."""
    participants = User.objects.filter(
        role=User.ROLE_PARTICIPANT,
    ).order_by('username')

    challenges = Challenge.objects.all().order_by('order')

    rows = []
    for user in participants:
        flags = {uf.challenge_id: uf.flag_value for uf in user.user_flags.all()}
        row_flags = [flags.get(ch.id, '—') for ch in challenges]
        rows.append({'user': user, 'flags': row_flags})

    return render(request, 'ctf/admin_user_flags.html', {
        'challenges': challenges,
        'rows': rows,
    })


@admin_required
@require_POST
def admin_regenerate_user_flags(request, user_id):
    """Regenerate flags for a specific user (e.g. after adding new challenges)."""
    user = get_object_or_404(User, pk=user_id)
    UserFlag.generate_flags_for_user(user)
    messages.success(request, f"Flags regenerated for '{user.username}'.")
    return redirect('admin_user_flags')


@admin_required
@require_POST
def remove_all_users(request):
    """Remove ALL non-admin users from the system."""
    users = User.objects.exclude(role=User.ROLE_ADMIN).exclude(is_superuser=True)
    count = users.count()
    # Delete related data first
    Score.objects.filter(user__in=users).delete()
    Submission.objects.filter(user__in=users).delete()
    UserFlag.objects.filter(user__in=users).delete()
    EvaluatorAssignment.objects.filter(
        Q(evaluator__in=users) | Q(participant__in=users)
    ).delete()
    users.delete()
    messages.success(request, f'All {count} users have been removed from the system.')
    return redirect('admin_dashboard')
