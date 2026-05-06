from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.utils import timezone
from django.contrib import messages
from .models import Event, Question, Participant, Answer
from .forms import ParticipantRegisterForm, LoginForm, EventSettingsForm
import json
import random
import logging

logger = logging.getLogger(__name__)


def home(request):
    if request.user.is_authenticated:
        if request.user.is_staff:
            return redirect('quiz:admin_dashboard')
        return redirect('quiz:waiting_room')
    return render(request, 'quiz/home.html')


def register(request):
    if request.method == 'POST':
        form = ParticipantRegisterForm(request.POST)
        if form.is_valid():
            user = form.save()
            event = Event.objects.first()
            Participant.objects.create(
                user=user,
                event=event,
                roll_no=form.cleaned_data.get('roll_no', ''),
                domain=form.cleaned_data.get('domain', ''),
            )
            messages.success(request, 'Registration successful! Please wait for admin approval.')
            return redirect('quiz:login')
    else:
        form = ParticipantRegisterForm()
    return render(request, 'quiz/register.html', {'form': form})


def login_view(request):
    if request.method == 'POST':
        form = LoginForm(request.POST)
        if form.is_valid():
            user = authenticate(
                request,
                username=form.cleaned_data['username'],
                password=form.cleaned_data['password']
            )
            if user is not None:
                login(request, user)
                if user.is_staff:
                    return redirect('quiz:admin_dashboard')
                return redirect('quiz:waiting_room')
            else:
                messages.error(request, 'Invalid credentials.')
    else:
        form = LoginForm()
    return render(request, 'quiz/login.html', {'form': form})


def logout_view(request):
    logout(request)
    return redirect('quiz:home')


@login_required
def waiting_room(request):
    if request.user.is_staff:
        return redirect('quiz:admin_dashboard')
    try:
        participant = request.user.r1_participant
    except Participant.DoesNotExist:
        event = Event.objects.first()
        participant = Participant.objects.create(user=request.user, event=event)

    event = participant.event
    context = {
        'participant': participant,
        'event': event,
    }

    if participant.status == 'disqualified':
        return render(request, 'quiz/disqualified.html', context)

    if participant.has_submitted:
        return redirect('quiz:result')

    if participant.status == 'approved' and event and event.is_running:
        return redirect('quiz:quiz')

    return render(request, 'quiz/waiting_room.html', context)


@login_required
def quiz_view(request):
    if request.user.is_staff:
        return redirect('quiz:admin_dashboard')
    try:
        participant = request.user.r1_participant
    except Participant.DoesNotExist:
        return redirect('quiz:waiting_room')
    event = participant.event

    if participant.status != 'approved':
        return redirect('quiz:waiting_room')

    if not event or not event.is_running:
        if not participant.has_submitted and participant.started_at:
            participant.finished_at = timezone.now()
            participant.has_submitted = True
            participant.calculate_score()
        return redirect('quiz:waiting_room')

    if participant.has_submitted:
        return redirect('quiz:result')

    if participant.status == 'disqualified':
        return render(request, 'quiz/disqualified.html')

    # Mark participant start time (their personal timer starts now)
    if not participant.started_at:
        participant.started_at = timezone.now()
        participant.save()

    # Auto-submit if participant's personal timer has expired
    if participant.is_time_up:
        participant.finished_at = timezone.now()
        participant.has_submitted = True
        participant.calculate_score()
        return redirect('quiz:result')

    questions = list(Question.objects.filter(event=event).order_by('question_no'))

    # Shuffle questions per user (consistent order using user ID as seed)
    rng = random.Random(request.user.id)
    rng.shuffle(questions)

    # Get existing answers
    answers = {}
    for ans in Answer.objects.filter(participant=participant, question__event=event):
        answers[ans.question_id] = ans.selected_option

    context = {
        'questions': questions,
        'answers_json': json.dumps(answers),
        'event': event,
        'participant': participant,
        'time_remaining': participant.time_remaining_seconds,
        'max_tab_switches': event.max_tab_switches,
        'tab_switch_count': participant.tab_switch_count,
    }
    return render(request, 'quiz/quiz.html', context)


@login_required
def save_answer(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        question_id = data.get('question_id')
        selected_option = data.get('selected_option')
        participant = request.user.r1_participant

        if participant.has_submitted or participant.status == 'disqualified':
            return JsonResponse({'status': 'error', 'message': 'Already submitted or disqualified'})

        question = get_object_or_404(Question, id=question_id)
        answer, created = Answer.objects.update_or_create(
            participant=participant,
            question=question,
            defaults={'selected_option': selected_option}
        )
        return JsonResponse({'status': 'ok'})
    return JsonResponse({'status': 'error'}, status=400)


@login_required
def record_tab_switch(request):
    if request.method == 'POST':
        participant = request.user.r1_participant
        event = participant.event

        if participant.has_submitted or participant.status == 'disqualified':
            return JsonResponse({'status': 'already_done'})

        participant.tab_switch_count += 1
        violated = participant.tab_switch_count >= event.max_tab_switches
        participant.tab_switch_violated = violated
        participant.save()

        if violated:
            # Auto-submit
            participant.finished_at = timezone.now()
            participant.has_submitted = True
            participant.calculate_score()
            return JsonResponse({
                'status': 'violated',
                'message': 'Tab switch limit exceeded. Your answers have been auto-submitted.',
                'count': participant.tab_switch_count,
            })

        return JsonResponse({
            'status': 'warning',
            'count': participant.tab_switch_count,
            'remaining': event.max_tab_switches - participant.tab_switch_count,
        })
    return JsonResponse({'status': 'error'}, status=400)


@login_required
def submit_quiz(request):
    if request.method == 'POST':
        participant = request.user.r1_participant
        if not participant.has_submitted:
            participant.finished_at = timezone.now()
            participant.has_submitted = True
            participant.calculate_score()
        return JsonResponse({'status': 'ok', 'score': participant.score})
    return JsonResponse({'status': 'error'}, status=400)


@login_required
def time_remaining(request):
    try:
        participant = request.user.r1_participant
    except Participant.DoesNotExist:
        return JsonResponse({'remaining': 0, 'is_running': False})
    event = participant.event
    if event:
        remaining = participant.time_remaining_seconds
        # Auto-submit server-side if time is up
        if participant.is_time_up and not participant.has_submitted and participant.started_at:
            participant.finished_at = timezone.now()
            participant.has_submitted = True
            participant.calculate_score()
            return JsonResponse({'remaining': 0, 'is_running': False, 'time_up': True})
        return JsonResponse({
            'remaining': remaining,
            'is_running': event.is_running,
        })
    return JsonResponse({'remaining': 0, 'is_running': False})


@login_required
def event_status(request):
    try:
        participant = request.user.r1_participant
    except Participant.DoesNotExist:
        return JsonResponse({'is_active': False, 'is_running': False, 'status': 'error'})
    event = participant.event
    if event:
        return JsonResponse({
            'is_active': event.is_active,
            'is_running': event.is_running,
            'status': participant.status,
            'has_submitted': participant.has_submitted,
        })
    return JsonResponse({'is_active': False, 'is_running': False, 'status': participant.status})


@login_required
def result_view(request):
    participant = request.user.r1_participant
    event = participant.event
    total = Question.objects.filter(event=event).count() if event else 0
    context = {
        'participant': participant,
        'total': total,
        'show_score': event.show_score_to_participant if event else False,
    }
    return render(request, 'quiz/result.html', context)


from quiz.constants import SECTION_RANGES, SECTION_TOTALS, SECTION_TOTALS_BY_DOMAIN, get_question_section, get_multiplier


def leaderboard(request):
    event = Event.objects.first()
    is_admin = request.user.is_authenticated and request.user.is_staff

    if event and not event.leaderboard_public and not is_admin:
        return render(request, 'quiz/leaderboard_private.html', {'event': event})

    participants = list(
        Participant.objects.filter(event=event, has_submitted=True)
        .exclude(status='disqualified')
        .order_by('-score', 'time_taken_ms')
        .select_related('user')
    )

    # Fetch all correct answers in one query
    all_answers = Answer.objects.filter(
        participant__in=participants,
        question__event=event,
    ).select_related('question')

    # Group answers by participant
    from collections import defaultdict
    p_answers = defaultdict(list)
    for ans in all_answers:
        p_answers[ans.participant_id].append(ans)

    # Build per-participant section scores (with cross-domain 2x multiplier)
    participants_data = []
    for p in participants:
        answers = p_answers[p.id]
        section_scores = {}
        for name, start, end in SECTION_RANGES:
            if start is None:
                section_scores[name] = 0
            else:
                section_scores[name] = sum(
                    get_multiplier(p.domain, name)
                    for a in answers
                    if start <= a.question.question_no <= end
                    and a.selected_option == a.question.correct_answer
                )
        participants_data.append({'participant': p, 'section_scores': section_scores})

    context = {
        'participants_data': participants_data,
        'sections': [s[0] for s in SECTION_RANGES],
        'section_totals': SECTION_TOTALS,
        'event': event,
        'is_admin': is_admin,
    }
    return render(request, 'quiz/leaderboard.html', context)


# ============ ADMIN VIEWS ============

@staff_member_required
def admin_dashboard(request):
    event = Event.objects.first()
    if not event:
        event = Event.objects.create(name="KIC AIML 2026 Assessment")

    participants = Participant.objects.filter(event=event).select_related('user')
    pending = participants.filter(status='pending')
    approved = participants.filter(status='approved')
    disqualified = participants.filter(status='disqualified')
    submitted = participants.filter(has_submitted=True)
    violated = participants.filter(tab_switch_violated=True)

    settings_form = EventSettingsForm(initial={
        'duration_minutes': event.duration_minutes,
        'max_tab_switches': event.max_tab_switches,
        'leaderboard_public': event.leaderboard_public,
        'show_score_to_participant': event.show_score_to_participant,
    })

    context = {
        'event': event,
        'participants': participants,
        'pending': pending,
        'approved': approved,
        'disqualified': disqualified,
        'submitted': submitted,
        'violated': violated,
        'settings_form': settings_form,
        'total_questions': Question.objects.filter(event=event).count(),
    }
    return render(request, 'quiz/admin_dashboard.html', context)


@staff_member_required
def approve_participant(request, participant_id):
    participant = get_object_or_404(Participant, id=participant_id)
    participant.status = 'approved'
    participant.save()
    messages.success(request, f'{participant.user.username} has been approved.')
    return redirect('quiz:admin_dashboard')


@staff_member_required
def approve_all(request):
    event = Event.objects.first()
    if event:
        Participant.objects.filter(event=event, status='pending').update(status='approved')
        messages.success(request, 'All pending participants have been approved.')
    return redirect('quiz:admin_dashboard')


@staff_member_required
def disqualify_participant(request, participant_id):
    participant = get_object_or_404(Participant, id=participant_id)
    participant.status = 'disqualified'
    if not participant.has_submitted and participant.started_at:
        participant.finished_at = timezone.now()
        participant.has_submitted = True
        participant.calculate_score()
    participant.save()
    messages.success(request, f'{participant.user.username} has been disqualified.')
    return redirect('quiz:admin_dashboard')


@staff_member_required
def start_event(request):
    event = Event.objects.first()
    if event:
        event.is_active = True
        if not event.started_at:
            event.started_at = timezone.now()
        event.save()
        messages.success(request, 'Event started! Each participant\'s timer begins when they enter the quiz.')
    return redirect('quiz:admin_dashboard')


@staff_member_required
def stop_event(request):
    event = Event.objects.first()
    if event:
        event.is_active = False
        event.save()
        # Auto-submit all who haven't submitted
        active = Participant.objects.filter(event=event, has_submitted=False, started_at__isnull=False)
        for p in active:
            p.finished_at = timezone.now()
            p.has_submitted = True
            p.calculate_score()
        messages.success(request, 'Event stopped! All pending submissions auto-submitted.')
    return redirect('quiz:admin_dashboard')


@staff_member_required
def update_settings(request):
    if request.method == 'POST':
        form = EventSettingsForm(request.POST)
        if form.is_valid():
            event = Event.objects.first()
            if event:
                event.duration_minutes = form.cleaned_data['duration_minutes']
                event.max_tab_switches = form.cleaned_data['max_tab_switches']
                event.leaderboard_public = form.cleaned_data['leaderboard_public']
                event.show_score_to_participant = form.cleaned_data['show_score_to_participant']
                event.save()
                messages.success(request, 'Settings updated!')
    return redirect('quiz:admin_dashboard')


@staff_member_required
def reset_quiz(request):
    """Reset quiz: clear all answers, scores, and submission state so users can take a new round."""
    event = Event.objects.first()
    if event:
        # Delete all answers
        Answer.objects.filter(participant__event=event).delete()

        # Reset all participant quiz state but keep their registration and approval
        Participant.objects.filter(event=event).update(
            started_at=None,
            finished_at=None,
            tab_switch_count=0,
            tab_switch_violated=False,
            score=0,
            time_taken_ms=0,
            has_submitted=False,
            status='pending',
        )

        # Reset event state — keep is_active=False so admin can start again
        event.is_active = False
        event.started_at = None
        event.save()

        messages.success(request, 'Quiz has been reset! All participants moved back to pending — approve them before starting.')
    return redirect('quiz:admin_dashboard')


@staff_member_required
def remove_participant(request, participant_id):
    """Remove a participant and their user account from the system entirely."""
    participant = get_object_or_404(Participant, id=participant_id)
    username = participant.user.username
    user = participant.user
    # Delete answers, participant, and the user account
    Answer.objects.filter(participant=participant).delete()
    participant.delete()
    user.delete()
    messages.success(request, f'{username} has been removed from the system.')
    return redirect('quiz:admin_dashboard')


@staff_member_required
def remove_all_users(request):
    """Remove ALL participants and their user accounts from the system."""
    event = Event.objects.first()
    if event:
        participants = Participant.objects.filter(event=event)
        count = participants.count()
        for p in participants:
            Answer.objects.filter(participant=p).delete()
            user = p.user
            p.delete()
            user.delete()
        messages.success(request, f'All {count} participants have been removed from the system.')
    return redirect('quiz:admin_dashboard')


@staff_member_required
def export_leaderboard_xlsx(request):
    """Export leaderboard results as XLSX — admin only."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    event = Event.objects.first()
    participants = Participant.objects.filter(
        event=event,
        has_submitted=True,
    ).exclude(status='disqualified').order_by('-score', 'time_taken_ms')

    total_questions = Question.objects.filter(event=event).count() if event else 0

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "KIC AIML 2026 - Leaderboard"

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    headers = [
        "Rank", "Name", "Email", "Roll No", "Domain",
        "PURE AI", "AI+DEV", "AI+CYBER",
        "PURE WEB", "WEB+CYBER", "WEB+DEV", "DBMS",
        "Final Score", "Time Taken", "Status",
        "NOTE: Cross-domain questions score 2x. DBMS is neutral (1x) for all domains."
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Fetch all answers for section scoring
    from collections import defaultdict
    all_answers = Answer.objects.filter(
        participant__in=participants, question__event=event
    ).select_related('question')
    p_answers = defaultdict(list)
    for ans in all_answers:
        p_answers[ans.participant_id].append(ans)

    for i, p in enumerate(participants, 1):
        total_sec = p.time_taken_ms // 1000
        mins, secs = divmod(total_sec, 60)
        time_str = f"{mins}m {secs}s"

        answers = p_answers[p.id]
        sec_scores = {}
        for name, start, end in SECTION_RANGES:
            if start is None:
                sec_scores[name] = 0
            else:
                sec_scores[name] = sum(
                    get_multiplier(p.domain, name)
                    for a in answers
                    if start <= a.question.question_no <= end
                    and a.selected_option == a.question.correct_answer
                )

        row = [
            i,
            f"{p.user.first_name} {p.user.last_name}".strip() or p.user.username,
            p.user.email,
            p.roll_no,
            p.domain,
            sec_scores['PURE AI'],
            sec_scores['AI+DEV'],
            sec_scores['AI+CYBER'],
            sec_scores['PURE WEB'],
            sec_scores['WEB+CYBER'],
            sec_scores['WEB+DEV'],
            sec_scores['DBMS'],
            p.score,
            time_str,
            p.get_status_display(),
        ]
        for col, val in enumerate(row, 1):
            cell = ws.cell(row=i + 1, column=col, value=val)
            cell.border = thin_border
            if col in (1, 13):
                cell.alignment = Alignment(horizontal="center")

    # Auto-width columns
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col) + 2
        ws.column_dimensions[col[0].column_letter].width = max_len

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="KIC_SE_Leaderboard.xlsx"'
    wb.save(response)
    return response
