from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, FileResponse, Http404, HttpResponse
from django.utils import timezone
from django.contrib import messages
from django.conf import settings as django_settings
from .models import Event, Challenge, Participant, Submission, Evaluation, CustomUser
from .forms import ParticipantRegisterForm, LoginForm, EventSettingsForm
import json
import os
from functools import wraps


def admin_required(view_func):
    """Like @staff_member_required but redirects to the normal login page."""
    @wraps(view_func)
    def _wrapped(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('debugchallenge:login')
        if not request.user.is_staff:
            return redirect('debugchallenge:home')
        return view_func(request, *args, **kwargs)
    return _wrapped


def home(request):
    if request.user.is_authenticated:
        if request.user.is_staff:
            return redirect('debugchallenge:admin_dashboard')
        try:
            if request.user.r2_participant.role == 'evaluator':
                return redirect('debugchallenge:evaluator_panel')
        except Participant.DoesNotExist:
            pass
        return redirect('debugchallenge:waiting_room')
    return render(request, 'debugchallenge/home.html')


def register(request):
    if request.method == 'POST':
        form = ParticipantRegisterForm(request.POST)
        if form.is_valid():
            user = form.save()
            event = Event.objects.first()
            Participant.objects.create(
                user=user,
                event=event,
                phone_number=form.cleaned_data.get('phone_number', ''),
            )
            messages.success(request, 'Registration successful! Please wait for admin approval.')
            return redirect('debugchallenge:login')
    else:
        form = ParticipantRegisterForm()
    return render(request, 'debugchallenge/register.html', {'form': form})


def login_view(request):
    if request.user.is_authenticated:
        if request.user.is_staff:
            return redirect('debugchallenge:admin_dashboard')
        return redirect('debugchallenge:waiting_room')

    if request.method == 'POST':
        form = LoginForm(request.POST)
        if form.is_valid():
            username_or_email = form.cleaned_data['username']
            password = form.cleaned_data['password']

            # Try authenticating with the value as username first
            user = authenticate(request, username=username_or_email, password=password)

            # If that fails, try looking up by email (for admin email login)
            if user is None:
                try:
                    lookup_user = CustomUser.objects.get(email=username_or_email)
                    user = authenticate(request, username=lookup_user.username, password=password)
                except CustomUser.DoesNotExist:
                    pass

            if user is not None:
                login(request, user)
                if user.is_staff:
                    return redirect('debugchallenge:admin_dashboard')
                try:
                    if user.r2_participant.role == 'evaluator':
                        return redirect('debugchallenge:evaluator_panel')
                except Participant.DoesNotExist:
                    pass
                return redirect('debugchallenge:waiting_room')
            else:
                messages.error(request, 'Invalid credentials.')
    else:
        form = LoginForm()
    return render(request, 'debugchallenge/login.html', {'form': form})


def logout_view(request):
    logout(request)
    return redirect('debugchallenge:home')


@login_required
def waiting_room(request):
    if request.user.is_staff:
        return redirect('debugchallenge:admin_dashboard')
    try:
        participant = request.user.r2_participant
    except Participant.DoesNotExist:
        event = Event.objects.first()
        participant = Participant.objects.create(user=request.user, event=event)

    # Evaluators go to their grading panel
    if participant.role == 'evaluator':
        return redirect('debugchallenge:evaluator_panel')

    event = participant.event
    context = {
        'participant': participant,
        'event': event,
    }

    if participant.status == 'disqualified':
        return render(request, 'debugchallenge/disqualified.html', context)

    if participant.has_submitted:
        return redirect('debugchallenge:result')

    # Approved participants always go to challenge page
    # (the challenge page itself handles active/inactive display)
    if participant.status == 'approved':
        return redirect('debugchallenge:challenge')

    return render(request, 'debugchallenge/waiting_room.html', context)


@login_required
def challenge_view(request):
    """Main challenge page — shows description + download when active, inactive message when not."""
    if request.user.is_staff:
        return redirect('debugchallenge:admin_dashboard')
    try:
        participant = request.user.r2_participant
    except Participant.DoesNotExist:
        return redirect('debugchallenge:waiting_room')
    event = participant.event

    if participant.status != 'approved':
        return redirect('debugchallenge:waiting_room')

    if participant.status == 'disqualified':
        return render(request, 'debugchallenge/disqualified.html', {
            'participant': participant,
            'event': event,
        })

    # Event active = visible to participants.  No timer / no tab-switch.
    challenges = Challenge.objects.filter(event=event).order_by('order') if event else Challenge.objects.none()

    context = {
        'challenges': challenges,
        'event': event,
        'participant': participant,
        'event_active': event.is_active if event else False,
    }
    return render(request, 'debugchallenge/challenge.html', context)


@login_required
def upload_submission(request):
    """Handle notebook file upload for a challenge."""
    if request.method == 'POST':
        participant = request.user.r2_participant

        if participant.has_submitted or participant.status == 'disqualified':
            return JsonResponse({'status': 'error', 'message': 'Already submitted or disqualified'})

        challenge_id = request.POST.get('challenge_id')
        uploaded_file = request.FILES.get('notebook_file')

        if not challenge_id or not uploaded_file:
            return JsonResponse({'status': 'error', 'message': 'Missing challenge_id or file'})

        # Validate file type
        if not uploaded_file.name.endswith('.ipynb'):
            return JsonResponse({'status': 'error', 'message': 'Only .ipynb files are accepted'})

        # Max file size: 10MB
        if uploaded_file.size > 10 * 1024 * 1024:
            return JsonResponse({'status': 'error', 'message': 'File too large. Max 10MB.'})

        challenge = get_object_or_404(Challenge, id=challenge_id)

        # Create or update submission
        submission, created = Submission.objects.update_or_create(
            participant=participant,
            challenge=challenge,
            defaults={'uploaded_file': uploaded_file}
        )

        return JsonResponse({
            'status': 'ok',
            'filename': os.path.basename(submission.uploaded_file.name),
            'submitted_at': submission.submitted_at.strftime('%H:%M:%S') if submission.submitted_at else '',
        })

    return JsonResponse({'status': 'error'}, status=400)


@login_required
def record_tab_switch(request):
    if request.method == 'POST':
        participant = request.user.r2_participant
        event = participant.event

        if participant.has_submitted or participant.status == 'disqualified':
            return JsonResponse({'status': 'already_done'})

        participant.tab_switch_count += 1
        violated = participant.tab_switch_count >= event.max_tab_switches
        participant.tab_switch_violated = violated
        participant.save()

        if violated:
            # Auto-submit
            participant.finished_at = timezone.now()
            participant.has_submitted = True
            participant.calculate_score()
            return JsonResponse({
                'status': 'violated',
                'message': 'Tab switch limit exceeded. Your submissions have been auto-finalized.',
                'count': participant.tab_switch_count,
            })

        return JsonResponse({
            'status': 'warning',
            'count': participant.tab_switch_count,
            'remaining': event.max_tab_switches - participant.tab_switch_count,
        })
    return JsonResponse({'status': 'error'}, status=400)


@login_required
def submit_challenge(request):
    """Finalize the challenge — no more uploads after this."""
    if request.method == 'POST':
        participant = request.user.r2_participant
        if not participant.has_submitted:
            participant.finished_at = timezone.now()
            participant.has_submitted = True
            participant.calculate_score()
        return JsonResponse({'status': 'ok', 'score': participant.score})
    return JsonResponse({'status': 'error'}, status=400)


@login_required
def time_remaining(request):
    try:
        participant = request.user.r2_participant
    except Participant.DoesNotExist:
        return JsonResponse({'remaining': 0, 'is_running': False})
    event = participant.event
    if event:
        remaining = participant.time_remaining_seconds
        # Auto-submit server-side if time is up
        if participant.is_time_up and not participant.has_submitted and participant.started_at:
            participant.finished_at = timezone.now()
            participant.has_submitted = True
            participant.calculate_score()
            return JsonResponse({'remaining': 0, 'is_running': False, 'time_up': True})
        return JsonResponse({
            'remaining': remaining,
            'is_running': event.is_running,
        })
    return JsonResponse({'remaining': 0, 'is_running': False})


@login_required
def event_status(request):
    try:
        participant = request.user.r2_participant
    except Participant.DoesNotExist:
        return JsonResponse({'is_active': False, 'is_running': False, 'status': 'error'})
    event = participant.event
    if event:
        return JsonResponse({
            'is_active': event.is_active,
            'is_running': event.is_running,
            'status': participant.status,
            'has_submitted': participant.has_submitted,
        })
    return JsonResponse({'is_active': False, 'is_running': False, 'status': participant.status})


@login_required
def result_view(request):
    participant = request.user.r2_participant
    event = participant.event
    challenges = Challenge.objects.filter(event=event) if event else Challenge.objects.none()
    submissions = Submission.objects.filter(participant=participant).select_related('challenge')
    total_bugs = sum(c.total_bugs for c in challenges)
    context = {
        'participant': participant,
        'total': total_bugs,
        'submissions': submissions,
        'challenges': challenges,
        'show_score': event.show_score_to_participant if event else False,
    }
    return render(request, 'debugchallenge/result.html', context)


def leaderboard(request):
    event = Event.objects.first()
    is_admin = request.user.is_authenticated and request.user.is_staff
    is_evaluator = False
    if request.user.is_authenticated and not request.user.is_staff:
        try:
            is_evaluator = request.user.r2_participant.role == 'evaluator'
        except Participant.DoesNotExist:
            pass

    # Enforce leaderboard_public for regular participants
    if not is_admin and not is_evaluator:
        if not event or not event.leaderboard_public:
            messages.info(request, 'The leaderboard is not available yet.')
            return redirect('debugchallenge:challenge')

    from django.db.models import Q
    participants = Participant.objects.filter(
        Q(event=event, role='participant'),
        Q(score__gt=0) | Q(has_submitted=True),
    ).exclude(status='disqualified').order_by('-score')

    total_bugs = sum(c.total_bugs for c in Challenge.objects.filter(event=event)) if event else 0

    # Admins/evaluators always see scores; participants only when toggle is on
    show_scores = is_admin or is_evaluator or (event and event.show_score_to_participant)

    context = {
        'participants': participants,
        'event': event,
        'is_admin': is_admin,
        'is_evaluator': is_evaluator,
        'total_bugs': total_bugs,
        'show_scores': show_scores,
    }
    return render(request, 'debugchallenge/leaderboard.html', context)


@login_required
def download_notebook(request, challenge_id):
    """Download a single challenge notebook."""
    try:
        participant = request.user.r2_participant
    except Participant.DoesNotExist:
        return redirect('debugchallenge:waiting_room')

    if participant.status != 'approved':
        return redirect('debugchallenge:waiting_room')

    event = participant.event
    if not event or not event.is_active:
        messages.warning(request, 'Challenge is not active yet.')
        return redirect('debugchallenge:challenge')

    challenge = get_object_or_404(Challenge, id=challenge_id, event=participant.event)
    materials_dir = os.path.join(django_settings.BASE_DIR, 'challenge_materials')
    notebook_path = os.path.join(materials_dir, challenge.notebook_filename)

    if not os.path.exists(notebook_path):
        raise Http404("Notebook file not found.")

    return FileResponse(
        open(notebook_path, 'rb'),
        content_type='application/octet-stream',
        as_attachment=True,
        filename=challenge.notebook_filename
    )


@login_required
def download_all_materials(request):
    """Download the pre-built challenge notebooks ZIP (notebooks + CSV)."""
    try:
        participant = request.user.r2_participant
    except Participant.DoesNotExist:
        return redirect('debugchallenge:waiting_room')

    if participant.status != 'approved':
        return redirect('debugchallenge:waiting_room')

    event = participant.event
    if not event or not event.is_active:
        messages.warning(request, 'Challenge is not active yet.')
        return redirect('debugchallenge:challenge')

    zip_path = os.path.join(
        django_settings.BASE_DIR, 'challenge_materials', 'ML_Challenge_Notebooks.zip'
    )
    if not os.path.exists(zip_path):
        raise Http404("Challenge notebooks ZIP not found.")

    return FileResponse(
        open(zip_path, 'rb'),
        content_type='application/zip',
        as_attachment=True,
        filename='ML_Challenge_Notebooks.zip'
    )


# ============ ADMIN VIEWS ============

@admin_required
def admin_dashboard(request):
    event = Event.objects.first()
    if not event:
        event = Event.objects.create(name="ML Fest Round 2 - ML Debugging Challenge")

    challenges = Challenge.objects.filter(event=event)
    participants = Participant.objects.filter(event=event).select_related('user')
    pending = participants.filter(status='pending')
    approved = participants.filter(status='approved')
    disqualified = participants.filter(status='disqualified')
    submitted = participants.filter(has_submitted=True)
    violated = participants.filter(tab_switch_violated=True)

    # Get submission counts per participant
    submission_data = {}
    for p in participants:
        subs = Submission.objects.filter(participant=p)
        submission_data[p.id] = {
            'count': subs.count(),
            'graded': subs.filter(is_graded=True).count(),
        }

    settings_form = EventSettingsForm(initial={
        'duration_minutes': event.duration_minutes,
        'max_tab_switches': event.max_tab_switches,
        'leaderboard_public': event.leaderboard_public,
        'show_score_to_participant': event.show_score_to_participant,
    })

    # Build evaluation leaderboard - participants ranked by evaluation score
    eval_participants = Participant.objects.filter(
        event=event, role='participant'
    ).select_related('user').order_by('-score', 'time_taken_ms')

    eval_leaderboard = []
    for p in eval_participants:
        cls_evals = Evaluation.objects.filter(participant=p, challenge_type='classification')
        reg_evals = Evaluation.objects.filter(participant=p, challenge_type='regression')
        cls_score = cls_evals.first().score if cls_evals.exists() else 0
        reg_score = reg_evals.first().score if reg_evals.exists() else 0
        total = cls_score + reg_score
        if total > 0 or cls_evals.exists() or reg_evals.exists():
            eval_leaderboard.append({
                'participant': p,
                'classification_score': cls_score,
                'regression_score': reg_score,
                'total_score': total,
                'graded': cls_evals.exists() or reg_evals.exists(),
            })
    eval_leaderboard.sort(key=lambda x: -x['total_score'])

    context = {
        'event': event,
        'challenges': challenges,
        'participants': participants,
        'pending': pending,
        'approved': approved,
        'disqualified': disqualified,
        'submitted': submitted,
        'violated': violated,
        'settings_form': settings_form,
        'total_challenges': challenges.count(),
        'submission_data': submission_data,
        'eval_leaderboard': eval_leaderboard,
        'classification_bugs': CLASSIFICATION_BUGS,
        'regression_bugs': REGRESSION_BUGS,
        'total_bugs': TOTAL_BUGS,
    }
    return render(request, 'debugchallenge/admin_dashboard.html', context)


@admin_required
def approve_participant(request, participant_id):
    participant = get_object_or_404(Participant, id=participant_id)
    participant.status = 'approved'
    participant.save()
    messages.success(request, f'{participant.user.username} has been approved.')
    return redirect('debugchallenge:admin_dashboard')


@admin_required
def approve_all(request):
    event = Event.objects.first()
    if event:
        Participant.objects.filter(event=event, status='pending').update(status='approved')
        messages.success(request, 'All pending participants have been approved.')
    return redirect('debugchallenge:admin_dashboard')


@admin_required
def disqualify_participant(request, participant_id):
    participant = get_object_or_404(Participant, id=participant_id)
    participant.status = 'disqualified'
    if not participant.has_submitted and participant.started_at:
        participant.finished_at = timezone.now()
        participant.has_submitted = True
        participant.calculate_score()
    participant.save()
    messages.success(request, f'{participant.user.username} has been disqualified.')
    return redirect('debugchallenge:admin_dashboard')


@admin_required
def start_event(request):
    event = Event.objects.first()
    if event:
        event.is_active = True
        event.save()
        messages.success(request, 'Event started! Challenge is now visible to participants.')
    return redirect('debugchallenge:admin_dashboard')


@admin_required
def stop_event(request):
    event = Event.objects.first()
    if event:
        event.is_active = False
        event.save()
        messages.success(request, 'Event stopped! Challenge is now hidden from participants.')
    return redirect('debugchallenge:admin_dashboard')


@admin_required
def update_settings(request):
    if request.method == 'POST':
        form = EventSettingsForm(request.POST)
        if form.is_valid():
            event = Event.objects.first()
            if event:
                event.duration_minutes = form.cleaned_data['duration_minutes']
                event.max_tab_switches = form.cleaned_data['max_tab_switches']
                event.leaderboard_public = form.cleaned_data['leaderboard_public']
                event.show_score_to_participant = form.cleaned_data['show_score_to_participant']
                event.save()
                messages.success(request, 'Settings updated!')
    return redirect('debugchallenge:admin_dashboard')


@admin_required
def reset_challenge(request):
    """Reset challenge: clear all submissions, scores, and submission state."""
    event = Event.objects.first()
    if event:
        # Delete all evaluations and submissions
        Evaluation.objects.filter(
            participant__event=event
        ).delete()
        Submission.objects.filter(participant__event=event).delete()

        # Reset all participant state but keep their registration and approval
        Participant.objects.filter(event=event).update(
            started_at=None,
            finished_at=None,
            tab_switch_count=0,
            tab_switch_violated=False,
            score=0,
            time_taken_ms=0,
            has_submitted=False,
            status='approved',
        )

        # Reset event state
        event.is_active = False
        event.started_at = None
        event.save()

        messages.success(request, 'Challenge has been reset! All registered participants are approved and ready.')
    return redirect('debugchallenge:admin_dashboard')


@admin_required
def remove_participant(request, participant_id):
    """Remove a participant and their user account from the system entirely."""
    try:
        participant = Participant.objects.get(id=participant_id)
    except Participant.DoesNotExist:
        messages.warning(request, 'Participant not found � may have already been removed.')
        return redirect('debugchallenge:admin_dashboard')
    username = participant.user.username
    user = participant.user
    Submission.objects.filter(participant=participant).delete()
    participant.delete()
    user.delete()
    messages.success(request, f'{username} has been removed from the system.')
    return redirect('debugchallenge:admin_dashboard')


@admin_required
def remove_all_users(request):
    """Remove ALL participants and their user accounts from the system."""
    event = Event.objects.first()
    if event:
        participants = Participant.objects.filter(event=event)
        count = participants.count()
        for p in participants:
            Evaluation.objects.filter(participant=p).delete()
            Evaluation.objects.filter(evaluator=p).delete()
            Submission.objects.filter(participant=p).delete()
            user = p.user
            p.delete()
            user.delete()
        messages.success(request, f'All {count} participants have been removed from the system.')
    return redirect('debugchallenge:admin_dashboard')


@admin_required
def grade_submission(request, submission_id):
    """Admin grades a submission — sets score."""
    if request.method == 'POST':
        submission = get_object_or_404(Submission, id=submission_id)
        try:
            score = int(request.POST.get('score', 0))
        except (ValueError, TypeError):
            score = 0
        notes = request.POST.get('admin_notes', '')
        submission.score = score
        submission.admin_notes = notes
        submission.is_graded = True
        submission.save()

        # Recalculate participant's total score
        submission.participant.calculate_score()

        messages.success(request, f'Graded {submission.participant.user.username} — {submission.challenge.title}: {score} pts')
    return redirect('debugchallenge:admin_dashboard')


@admin_required
def download_submission(request, submission_id):
    """Admin downloads a participant's submitted notebook."""
    submission = get_object_or_404(Submission, id=submission_id)
    if not submission.uploaded_file:
        raise Http404("No file uploaded.")

    filename = f"{submission.participant.user.username}_{submission.challenge.challenge_type}.ipynb"
    return FileResponse(
        submission.uploaded_file.open('rb'),
        content_type='application/octet-stream',
        as_attachment=True,
        filename=filename
    )


@admin_required
def view_submissions(request, participant_id):
    """View all submissions for a specific participant."""
    participant = get_object_or_404(Participant, id=participant_id)
    submissions = Submission.objects.filter(participant=participant).select_related('challenge')
    challenges = Challenge.objects.filter(event=participant.event)

    context = {
        'participant': participant,
        'submissions': submissions,
        'challenges': challenges,
    }
    return render(request, 'debugchallenge/view_submissions.html', context)


# ============ EVALUATOR VIEWS ============

@admin_required
def set_role(request, participant_id):
    """Admin sets a participant's role to 'participant' or 'evaluator'."""
    if request.method == 'POST':
        participant = get_object_or_404(Participant, id=participant_id)
        new_role = request.POST.get('role', 'participant')
        if new_role in ('participant', 'evaluator'):
            old_role = participant.role
            participant.role = new_role
            # If changing to evaluator, clear their own evaluator assignment
            if new_role == 'evaluator':
                participant.evaluator = None
            # If changing from evaluator to participant, unassign all their assigned participants
            if old_role == 'evaluator' and new_role == 'participant':
                Participant.objects.filter(evaluator=participant).update(evaluator=None)
            participant.save()
            messages.success(request, f'{participant.user.username} is now an {new_role}.')
    return redirect('debugchallenge:admin_dashboard')


@admin_required
def evaluator_management(request):
    """Evaluator management page: list evaluators and their assigned participants."""
    event = Event.objects.first()
    evaluators = Participant.objects.filter(
        event=event, role='evaluator'
    ).select_related('user')

    # Build assignment data for each evaluator
    evaluator_data = []
    for ev in evaluators:
        assigned = Participant.objects.filter(
            evaluator=ev, role='participant'
        ).select_related('user')
        evaluator_data.append({
            'evaluator': ev,
            'assigned': assigned,
            'assigned_count': assigned.count(),
        })

    # Unassigned participants (role=participant, no evaluator assigned)
    unassigned = Participant.objects.filter(
        event=event, role='participant', evaluator__isnull=True
    ).select_related('user')

    # All participants available for assignment
    all_participants = Participant.objects.filter(
        event=event, role='participant'
    ).select_related('user')

    assigned_count = all_participants.exclude(evaluator__isnull=True).count()

    # Build evaluation leaderboard - participants ranked by evaluation score (descending)
    eval_participants = Participant.objects.filter(
        event=event, role='participant'
    ).select_related('user').order_by('-score', 'time_taken_ms')

    eval_leaderboard = []
    for p in eval_participants:
        cls_evals = Evaluation.objects.filter(participant=p, challenge_type='classification')
        reg_evals = Evaluation.objects.filter(participant=p, challenge_type='regression')
        cls_score = cls_evals.first().score if cls_evals.exists() else 0
        reg_score = reg_evals.first().score if reg_evals.exists() else 0
        total = cls_score + reg_score
        if total > 0 or cls_evals.exists() or reg_evals.exists():
            eval_leaderboard.append({
                'participant': p,
                'classification_score': cls_score,
                'regression_score': reg_score,
                'total_score': total,
                'graded': cls_evals.exists() or reg_evals.exists(),
            })
    eval_leaderboard.sort(key=lambda x: -x['total_score'])

    context = {
        'event': event,
        'evaluator_data': evaluator_data,
        'evaluators': evaluators,
        'unassigned': unassigned,
        'all_participants': all_participants,
        'assigned_count': assigned_count,
        'eval_leaderboard': eval_leaderboard,
        'classification_bugs': CLASSIFICATION_BUGS,
        'regression_bugs': REGRESSION_BUGS,
        'total_bugs': TOTAL_BUGS,
    }
    return render(request, 'debugchallenge/evaluator_management.html', context)


@admin_required
def assign_participants(request):
    """Assign multiple participants to an evaluator."""
    if request.method == 'POST':
        evaluator_id = request.POST.get('evaluator_id')
        participant_ids = request.POST.getlist('participant_ids')

        evaluator = get_object_or_404(Participant, id=evaluator_id, role='evaluator')

        # Clear previous assignments for this evaluator
        Participant.objects.filter(evaluator=evaluator).update(evaluator=None)

        # Assign selected participants
        if participant_ids:
            Participant.objects.filter(
                id__in=participant_ids, role='participant'
            ).update(evaluator=evaluator)

        count = len(participant_ids)
        messages.success(request, f'{count} participant(s) assigned to evaluator @{evaluator.user.username}.')

    return redirect('debugchallenge:evaluator_management')


# ============ EVALUATOR PANEL VIEWS ============

CLASSIFICATION_BUGS = 12
REGRESSION_BUGS = 6
TOTAL_BUGS = CLASSIFICATION_BUGS + REGRESSION_BUGS


@login_required
def evaluator_panel(request):
    """Evaluator's grading panel - shows assigned participants and scores."""
    try:
        evaluator = request.user.r2_participant
    except Participant.DoesNotExist:
        return redirect('debugchallenge:home')

    if evaluator.role != 'evaluator':
        return redirect('debugchallenge:waiting_room')

    assigned = Participant.objects.filter(
        evaluator=evaluator, role='participant'
    ).select_related('user').order_by('user__username')

    participant_data = []
    for p in assigned:
        cls_eval = Evaluation.objects.filter(
            evaluator=evaluator, participant=p, challenge_type='classification'
        ).first()
        reg_eval = Evaluation.objects.filter(
            evaluator=evaluator, participant=p, challenge_type='regression'
        ).first()
        cls_score = cls_eval.score if cls_eval else 0
        reg_score = reg_eval.score if reg_eval else 0
        participant_data.append({
            'participant': p,
            'classification_score': cls_score,
            'classification_graded': cls_eval is not None,
            'regression_score': reg_score,
            'regression_graded': reg_eval is not None,
            'total_score': cls_score + reg_score,
        })

    context = {
        'evaluator': evaluator,
        'participant_data': participant_data,
        'classification_bugs': CLASSIFICATION_BUGS,
        'regression_bugs': REGRESSION_BUGS,
        'total_bugs': TOTAL_BUGS,
    }
    return render(request, 'debugchallenge/evaluator_panel.html', context)


@login_required
def evaluate_participant(request, participant_id):
    """Evaluator grades a participant's bugs - classification (12) + regression (6)."""
    try:
        evaluator = request.user.r2_participant
    except Participant.DoesNotExist:
        return redirect('debugchallenge:home')

    if evaluator.role != 'evaluator':
        return redirect('debugchallenge:waiting_room')

    participant = get_object_or_404(
        Participant, id=participant_id, evaluator=evaluator, role='participant'
    )

    # Get or create evaluation records
    cls_eval, _ = Evaluation.objects.get_or_create(
        evaluator=evaluator, participant=participant, challenge_type='classification',
        defaults={'bug_results': {str(i): False for i in range(1, CLASSIFICATION_BUGS + 1)}}
    )
    reg_eval, _ = Evaluation.objects.get_or_create(
        evaluator=evaluator, participant=participant, challenge_type='regression',
        defaults={'bug_results': {str(i): False for i in range(1, REGRESSION_BUGS + 1)}}
    )

    if request.method == 'POST':
        # Process classification bugs
        cls_results = {}
        for i in range(1, CLASSIFICATION_BUGS + 1):
            cls_results[str(i)] = request.POST.get(f'cls_bug_{i}') == 'on'
        cls_eval.bug_results = cls_results
        cls_eval.score = sum(1 for v in cls_results.values() if v)
        cls_eval.save()

        # Process regression bugs
        reg_results = {}
        for i in range(1, REGRESSION_BUGS + 1):
            reg_results[str(i)] = request.POST.get(f'reg_bug_{i}') == 'on'
        reg_eval.bug_results = reg_results
        reg_eval.score = sum(1 for v in reg_results.values() if v)
        reg_eval.save()

        # Update participant's total score
        participant.calculate_score()

        messages.success(
            request,
            f'Evaluation saved for @{participant.user.username}. '
            f'Classification: {cls_eval.score}/{CLASSIFICATION_BUGS}, '
            f'Regression: {reg_eval.score}/{REGRESSION_BUGS}, '
            f'Total: {participant.score}/{TOTAL_BUGS}'
        )
        return redirect('debugchallenge:evaluator_panel')

    # Build bug lists with current state
    cls_bugs = []
    for i in range(1, CLASSIFICATION_BUGS + 1):
        cls_bugs.append({
            'number': i,
            'checked': cls_eval.bug_results.get(str(i), False),
        })
    reg_bugs = []
    for i in range(1, REGRESSION_BUGS + 1):
        reg_bugs.append({
            'number': i,
            'checked': reg_eval.bug_results.get(str(i), False),
        })

    context = {
        'evaluator': evaluator,
        'participant': participant,
        'cls_eval': cls_eval,
        'reg_eval': reg_eval,
        'cls_bugs': cls_bugs,
        'reg_bugs': reg_bugs,
        'classification_bugs': CLASSIFICATION_BUGS,
        'regression_bugs': REGRESSION_BUGS,
        'total_bugs': TOTAL_BUGS,
    }
    return render(request, 'debugchallenge/evaluate_participant.html', context)


@login_required
def export_leaderboard_xlsx(request):
    """Export leaderboard results as XLSX — admin and evaluators only."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    is_admin = request.user.is_staff
    is_evaluator = False
    if not is_admin:
        try:
            is_evaluator = request.user.r2_participant.role == 'evaluator'
        except Participant.DoesNotExist:
            pass
    if not is_admin and not is_evaluator:
        return redirect('debugchallenge:leaderboard')

    event = Event.objects.first()
    from django.db.models import Q
    participants = Participant.objects.filter(
        Q(event=event, role='participant'),
        Q(score__gt=0) | Q(has_submitted=True),
    ).exclude(status='disqualified').order_by('-score')

    total_bugs = sum(c.total_bugs for c in Challenge.objects.filter(event=event)) if event else 0

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Round 2 - Leaderboard"

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    headers = ["Rank", "Username", "Email", "Phone", "Score", "Total Bugs", "Status"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for i, p in enumerate(participants, 1):
        row = [
            i,
            p.user.username,
            p.user.email,
            p.phone_number,
            p.score,
            total_bugs,
            p.get_status_display(),
        ]
        for col, val in enumerate(row, 1):
            cell = ws.cell(row=i + 1, column=col, value=val)
            cell.border = thin_border
            if col in (1, 5, 6):
                cell.alignment = Alignment(horizontal="center")

    # Auto-width columns
    for col_cells in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col_cells) + 2
        ws.column_dimensions[col_cells[0].column_letter].width = max_len

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="Round2_Leaderboard.xlsx"'
    wb.save(response)
    return response
